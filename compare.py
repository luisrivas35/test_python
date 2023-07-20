#!/usr/bin/env python3
#author:Luis Rivas
#script to compare two .txt files and find the no match components

from openpyxl import Workbook
from openpyxl import load_workbook

wb = openpyxl.load_workbook("C:\\test\\compare.xlsx")
sh1 = wb['PIM']
sh2 = wb['PR']
row = sh1.max_row
column = sh1.max_column

for i in range(1,column+1)